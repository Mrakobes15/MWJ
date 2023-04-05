import requests
from bs4 import BeautifulSoup as Bs
import datetime
from openpyxl import Workbook
from os import path, mkdir
import numpy as np
import matplotlib.pyplot as plt
from progress.bar import Bar





def get_page(page_utl: str):
    request = requests.get(page_utl)
    page = Bs(request.content, 'html.parser')
    return page


def microwavejournal(older_date=datetime.date.today() - datetime.timedelta(30)):
    months_dict = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                   'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12}

    page_num = 1
    article_num = 1
    articles_info = list()
    tag_cunt = dict()
    while True:
        page = get_page(page_utl='https://www.microwavejournal.com/articles/topic/3372?page=' + str(page_num))
        articles_list = page.select('.article-summary__details')
        if len(articles_list) > 0:
            x = 1
            for article in articles_list:
                x += 1
                info_dict = dict()
                class_dick = {'Дата': '.date', 'Заголовок': '.headline > a', 'Описание': '.abstract > p'}
                for class_key in class_dick.keys():
                    try:
                        select = article.select_one(class_dick[class_key])
                        info = select.text
                        if class_key == 'Дата':
                            month, day, year = info.split()
                            info = '{year}.{month}.{day}'.format(
                                year=year, month=months_dict[month], day=day[:-1]
                            )
                            if datetime.datetime.strptime(info, '%Y.%m.%d').date() < older_date:
                                for i in tag_cunt:
                                    print(f'{i}: {tag_cunt[i]}')
                                return articles_info
                        if class_key == 'Заголовок':
                            tag_list = [tag.text for tag in get_page(select['href']).select('.tags > a')]
                            for tag in tag_list:
                                if tag in tag_cunt.keys():
                                    tag_cunt[tag] += 1
                                else:
                                    tag_cunt[tag] = 1
                            info_dict['Теги'] = ', '.join(tag_list)
                        info_dict[class_key] = info
                    except AttributeError:
                        info_dict[class_key] = 'Нет данных'
                articles_info.append(info_dict)
                print('Загруженно статей: {article_num}'.format(article_num=article_num))
                article_num += 1
        else:
            for i in tag_cunt:
                print(f'{i}: {tag_cunt[i]}')
            return articles_info
        page_num += 1


def normalization_date(date_text: str):
    try:
        normalized_date = datetime.datetime.strptime(date_text, '%Y.%m.%d').date()
        return normalized_date
    except ValueError:
        input("Некорректная дата! Введите дату ещё раз: ")


def exel_maker(export_date: list[dict]):
    book = Workbook()
    sheet = book.active
    row, column = 1, 1
    for article in export_date:
        for title in article:
            if row == 1:
                sheet.cell(row=row, column=column).value = title
            else:
                sheet.cell(row=row, column=column).value = article[title]
            column += 1
        column = 1
        row += 1

    if not(path.exists('tables')):
        mkdir('tables')
    table_name = 'tables/' + input('Введите имя таблицы: ') + '.xlsx'
    if path.exists(table_name):
        name_flag = True
    else:
        book.save(table_name)
        name_flag = False
    while name_flag:
        answer = input('Такой файл уже существует!\n'
                           '1. Переименовать\n'
                           '2. Перезаписать\n'
                           '3. Закончить\n'
                           'Введите номер команды: ')
        match answer:
            case '1':
                table_name = 'tables/' + input('Введите имя таблицы: ') + '.xlsx'
            case '2':
                book.save(table_name)
                name_flag = False
            case '3':
                name_flag = False
            case _:
                print('Такой команды нет')

    book.close()


if __name__ == '__main__':
    print('До какой даты смотерть статьи?')
    end_date = normalization_date(date_text=input('Введите дату (гггг.мм.дд): '))
    print('Загрузка...')
    news = microwavejournal(older_date=end_date)
    exel_maker(export_date=news)

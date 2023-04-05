from openpyxl import Workbook
from wordcloud import WordCloud, STOPWORDS
from stop_words import get_stop_words
import matplotlib.pyplot as plt
import openpyxl
import re
import sqlite3


def get_data(name: str, start: int, end: int):
    wb = openpyxl.load_workbook("tables/" + name + ".xlsx")
    ws = wb.active
    text = ''
    for i in range(1, ws.max_row):
        for col in ws.iter_cols(start, end):
            if col[i].value is not None:
                text += '\n'
                text += col[i].value

    return text


def get_companies():
    wb = openpyxl.load_workbook("tables/companies.xlsx")
    ws = wb.active
    text = []
    for i in range(1, ws.max_row):
        for col in ws.iter_cols(1, 1):
            if col[i].value is not None:
                text.append(col[i].value)

    return text


def get_cloud():
    STOPWORDS_EN = get_stop_words("english")
    data = get_data()
    data = re.sub(r'==.*?|==+', '', data)
    STOPWORDS_EN.append('Нет')
    STOPWORDS_EN.append('данных')
    STOPWORDS_EN.append('will')
    STOPWORDS_EN.append('can')
    wordcloud = WordCloud(width=2000,
                          height=1500,
                          max_words=150,
                          random_state=1,
                          background_color='black',
                          margin=20,
                          colormap='Pastel1',
                          collocations=False,
                          stopwords=STOPWORDS_EN).generate(data)
    wordcloud.to_file("wc.png")


def push_companies_db(companies: dict):
    db_connection = sqlite3.connect('MWJ.db')
    cursor = db_connection.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS companies(id INT PRIMARY KEY, company TEXT, mentioned TEXT);")
    data1 = []
    keys = list(companies.keys())
    values = list(companies.values())
    for i in range(1, len(companies)):
        data1.append(int(i))
        data1.append(keys[i])
        data1.append(values[i])
        t = tuple(data1)
        cursor.execute("INSERT INTO companies VALUES(?, ?, ?);", t)
        data1.clear()
    db_connection.commit()
    db_connection.close()


#get_cloud()

def companies_mentioned(tag_count: dict, data: list, company: list):
    if len(tag_count) == 0:
        for tag in company:
            if tag in tag_count.keys():
                tag_count[tag] += 1
            else:
                tag_count[tag] = 1

    for tag in tag_count.keys():
        for words in data:
            if words.find(tag) != -1:
                tag_count[tag] += 1


def get_column(column: str, table_name: str):
    db_connection = sqlite3.connect('MWJ.db')
    cursor = db_connection.cursor()
    cursor.execute("SELECT " + column + " FROM " + table_name)
    radar_tags = cursor.fetchall()
    text = ''
    for tag in radar_tags:
        for tup in tag:
            text += tup
            text += ','
    radar_tags.clear()
    radar_tags = text.split(',')
    return radar_tags





#dat = get_data("30articles", 2, 3).split('\n')
#comp = get_column('company', 'companies')

tags = dict()
#companies_mentioned(tags, dat, comp)
#print(len(tags), tags)




#RefreshMyTags("AerospaceDefense")
#companies_mentioned(tags, dat, comp)
#print(len(tags), tags)
#push_companies_db(tags)


#"CREATE TABLE IF NOT EXISTS users(id INT PRIMARY KEY, company TEXT, mentioned TEXT);"
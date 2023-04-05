import sqlite3
import openpyxl


def push_data_xlsx_db(name: str):
    wb = openpyxl.load_workbook("tables/" + name)
    ws = wb.active
    data = []
    table_name = input("Table name: ")
    for i in range(1, ws.max_row):
        data.append(int(i))
        for col in ws.iter_cols(1, 4):
            data.append(col[i].value)
        t = tuple(data)
        cursor.execute("INSERT INTO " + table_name + " VALUES(?, ?, ?, ?, ?);", t)
        data.clear()


try:
    db_connection = sqlite3.connect('MWJ.db')
    cursor = db_connection.cursor()
    sqlite_select_query = "select sqlite_version();"
    cursor.execute(sqlite_select_query)
    record = cursor.fetchall()
    print("Database version: ", record)
    script = input(" Enter sqlite query")
    if script is not None:
        result = cursor.execute(script).fetchall()
    if result is not None:
        print(result)
    import_data = input("Import data?\n 1. Yes\n 2. No\n")
    if import_data == '1':
        push_data_xlsx_db(input("Xlsx filename: "))
    db_connection.commit()
except sqlite3.Error as error:
    print("Connection failed: ", error)
finally:
    if (db_connection):
        db_connection.close()
        print("Connection closed.")
